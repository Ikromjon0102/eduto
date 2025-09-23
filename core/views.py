import json
from django.contrib import messages
from django.contrib.auth.views import LoginView
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import DetailView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from datetime import date
import openpyxl
from django.http import HttpResponse, JsonResponse
from reportlab.pdfgen import canvas
from django.views import View
from .models import *
from .forms import (UserCreateForm, CourseCreateForm,
                    GroupCreateForm, PaymentPersonalForm,
                    PaymentForm, GradeForm, ApplicantCreateForm)
from .models import User, Group, Course
from django.views.generic import ListView
from django.utils import timezone
from django.db.models import Subquery, OuterRef, Q
from datetime import datetime
from django.db.models import Sum, F, Count
from django.contrib.auth.decorators import login_required, user_passes_test



class CustomLoginView(LoginView):
    template_name = 'login.html'  # HTML fayl nomi
    redirect_authenticated_user = True  # Foydalanuvchi login bo'lsa, yo'naltirish
    success_url = None  # success_url ni dinamik belgilaymiz

    def get_success_url(self):
        # Foydalanuvchi roli asosida URLni tanlaymiz
        if self.request.user.role == 'teacher':
            return reverse_lazy('teacher_profile')
        return reverse_lazy('manager_page')


# class Profile(View):
#
#     def get(self, request):
#         form1 = UserCreateForm()
#         form2 = CourseCreateForm()
#         form3 = GroupCreateForm()
#         students = User.objects.filter(role = 'student')
#         teachers = User.objects.filter(role = 'teacher')
#         courses = Course.objects.all()
#         groups = Group.objects.all()
#
#         context = {
#             'courses': courses,
#             'groups': groups,
#             'form_student': form1,
#             'form_course': form2,
#             'form_group': form3,
#             'teachers':teachers,
#             'students': students
#         }
#         return render(request, 'admin.html', context )
#
#
#     def post(self, request):
#         create_form = UserCreateForm(data=request.POST)
#         create_course_form = CourseCreateForm(data=request.POST)
#         create_group_form = GroupCreateForm(data=request.POST)
#         if create_form.is_valid():
#             create_form.save()
#             return redirect('manager_page')
#
#         elif create_course_form.is_valid():
#             create_course_form.save()
#             return redirect('manager_page')
#
#         elif create_group_form.is_valid():
#             create_group_form.save()
#             return redirect('manager_page')
#         else:
#             context = {
#                 'form_student': create_form,
#                 'form_course': create_course_form,
#                 'form_group': create_group_form,
#             }
#             return render(request, 'admin.html', context)

class Profile(View):
    def get(self, request):
        form1 = ApplicantCreateForm()
        form2 = CourseCreateForm()
        form3 = GroupCreateForm()
        form4 = UserCreateForm()
        students = User.objects.filter(Q(role='student') & Q(group__isnull=False))
        applicants = User.objects.filter(role='applicant')# Applicantlar
        # applicants = User.objects.filter(Q(role='applicant') | Q(role='student',  group=None)) # Applicantlar
        teachers = User.objects.filter(role='teacher')
        courses = Course.objects.all()
        groups = Group.objects.all()

        context = {
            'courses': courses,
            'groups': groups,
            'form_student': form1,
            'form_course': form2,
            'form_group': form3,
            'form_teacher': form4,
            'teachers': teachers,
            'students': students,
            'applicants': applicants,
        }
        return render(request, 'admin.html', context)

    def post(self, request):
        create_form = UserCreateForm(data=request.POST)
        create_course_form = CourseCreateForm(data=request.POST)
        create_group_form = GroupCreateForm(data=request.POST)
        applicant_form = ApplicantCreateForm(data=request.POST)  # Yangi qo‘shildi

        if applicant_form.is_valid():
            applicant_form.save()
            return redirect('manager_page')

        elif create_form.is_valid():
            create_form.save()
            return redirect('manager_page')

        elif create_course_form.is_valid():
            create_course_form.save()
            return redirect('manager_page')

        elif create_group_form.is_valid():
            create_group_form.save()
            return redirect('manager_page')

        else:
            context = {
                'form_student': applicant_form,  # O'zgartirildi
                'form_course': create_course_form,
                'form_group': create_group_form,
                'form_teacher': create_form,  # O'zgartirildi
            }
            return render(request, 'admin.html', context)


class UstozPorfileView(View):
    def get(self, request, ustoz_id):
        ustoz = User.objects.get(id=ustoz_id)
        groups = Group.objects.filter(teacher=ustoz)

        # Har bir guruhda o'quvchilarni filtrlash (faqat studentlar)
        groups_with_students = []
        alll = 0
        for group in groups:
            students = group.users.filter(role='student')
            alll += len(students)
            groups_with_students.append({'group': group, 'students': students, 'count':len(students)})

        context = {
            'teacher': ustoz,
            'groups_with_students': groups_with_students,
            'total_students': alll
        }
        return render(request, 'teacher_page.html', context)


def generate_pdf_report(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="student_report.pdf"'

    # PDF fayl yaratish
    p = canvas.Canvas(response)
    p.drawString(100, 800, "Guruh va O'quvchilar Hisoboti")

    students = User.objects.filter(role='student')
    y = 750
    for student in students:
        p.drawString(100, y, f"{student.username} - {student.group}")
        y -= 20

    p.showPage()
    p.save()
    return response


def generate_excel_report(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="student_report.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "O'quvchilar Hisoboti"

    # Sarlavhalar
    sheet.append(["Username", "Ism", "Guruh", ])

    students = User.objects.filter(role='student')
    for student in students:
        # student.group.name orqali guruhning nomini olish
        group_name = student.group.name if student.group else "Noma'lum"
        sheet.append([student.username, student.first_name, group_name])

    workbook.save(response)
    return response


def download_template_with_data(request):
    # Fayl va sheet yaratish
    workbook = openpyxl.Workbook()
    all_sheet = workbook.active
    all_sheet.title = 'all'

    # First sheet uchun template qo'shish
    all_sheet.append(['first_name', 'last_name', 'course_id', 'group_id', 'ustoz_id', 'phone'])

    id_lar_sheet = workbook.create_sheet(title='id_lar')

    # Ustozlar ro'yxatini yozish
    teachers = User.objects.filter(role='teacher')
    id_lar_sheet.append(['Ustoz ID', 'Ustoz Ismi'])
    for teacher in teachers:
        id_lar_sheet.append([teacher.id, teacher.username])

    # Guruhlar ro'yxatini yozish
    id_lar_sheet.append([])  # Bo'sh qator
    id_lar_sheet.append(['Guruh ID', 'Guruh Nomi'])
    groups = Group.objects.all()
    for group in groups:
        id_lar_sheet.append([group.id, group.name])

    # Kurslar ro'yxatini yozish
    id_lar_sheet.append([])  # Bo'sh qator
    id_lar_sheet.append(['Kurs ID', 'Kurs Nomi'])
    courses = Course.objects.all()
    for course in courses:
        id_lar_sheet.append([course.id, course.name])

    # Javob qaytarish
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=import_template.xlsx'

    workbook.save(response)
    return response

def import_from_excel(request):

    # print(request)
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']

        # Excel faylini o'qish
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb['jami']  # Bu yerda 'jami' sheet nomi bo'lishi kerak

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Min row 2dan boshlab ma'lumotni o'qiydi
            first_name, last_name, course_id, group_id, teacher_id, phone = row

            # Kursni ID orqali olish
            try:
                course = Course.objects.get(id=course_id)
            except Course.DoesNotExist:
                course = None  # Agar kurs topilmasa, uni None qilib qo'yish

            # Guruhni ID orqali olish
            try:
                group = Group.objects.get(id=group_id)
            except Group.DoesNotExist:
                group = None  # Agar guruh topilmasa, uni None qilib qo'yish

            # Ustozni ID orqali olish
            try:
                teacher = User.objects.get(id=teacher_id, role='teacher')
            except User.DoesNotExist:
                teacher = None  # Agar ustoz topilmasa, uni None qilib qo'yish

            # User modelini yaratish
            if course and group and teacher:  # Agar barcha bog'lanishlar to'g'ri bo'lsa
                user = User.objects.create(
                    username = f"{first_name.lower()}_{last_name.lower()}",
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone,
                    role='student',
                    group=group
                )
                user.save()
            else:
                # Agar biron narsa noto'g'ri bo'lsa, xabarni chiqaring
                print(f"User for {first_name} {last_name} not created due to missing relationships")

        # Import jarayoni tugagach, foydalanuvchini manager_page sahifasiga yo'naltirish
        return redirect('manager_page')  # 'manager_page' URL to'g'ri ekanligiga ishonch hosil qiling

    return render(request, 'upload.html')  # Agar POST bo'lmasa, formani ko'rsatish


def excel_page(request):
    return render(request, 'payment.html')


class PaymentView(ListView):
    model = Payment
    template_name = 'payment.html'
    context_object_name = 'payments'

    def get_queryset(self):
        return Payment.objects.all().order_by('student', 'month_period')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        payment_status = []
        for payment in context['payments']:
            deadline = payment.month_period.end_date
            is_overdue = payment.status == 'unpaid' and today > deadline
            payment_status.append({'payment': payment, 'is_overdue': is_overdue})
        context['payment_status'] = payment_status
        return context


def load_students(request):
    group_id = request.GET.get('group')
    if group_id:
        students = User.objects.filter(group_id=group_id, role='student')
        return JsonResponse({
            'students': [
                {'id': s.id, 'name': f"{s.first_name} {s.last_name}"}
                for s in students
            ]
        })
    return JsonResponse({'students': []})

# def load_students(request):
#     """Guruh bo'yicha o'quvchilarni yuklash uchun AJAX view"""
#     group_id = request.GET.get('group')
#     if group_id:
#         students = User.objects.filter(group_id=group_id, role='student')
#         print(group_id, students)
#         return JsonResponse({
#             'students': [{'id': student.id, 'name': student.get_full_name() or student.username}
#                          for student in students]
#         })
#
#     return JsonResponse({'students': []})


def group_details(request, group_id):
    """Guruh va kurs ma'lumotlarini olish uchun API view"""
    try:
        group = Group.objects.get(id=group_id)
        data = {
            'course': {
                'id': group.course.id,
                'price_per_month': float(group.course.price_per_month)
            }
        }
        return JsonResponse(data)
    except Group.DoesNotExist:
        return JsonResponse({'error': 'Group not found'}, status=404)


class StudentProfileView(LoginRequiredMixin, DetailView):
    model = User
    template_name = 'students/profile.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object
        if student.group:
            context['course'] = student.group.course
            context['start_date'] = student.group.start_date
            context['end_date'] = student.group.end_date

            start_date = student.group.start_date
            today = date.today()
            months_studied = relativedelta(today, start_date).months + 1
            context['current_month'] = min(months_studied, student.group.course.duration_months)

            # To'lovlarni oylik davr bo'yicha olish
            payments = Payment.objects.filter(student=student, group=student.group)
            all_payments = []
            # for month_period in MonthPeriod.objects.filter(is_active=True):
            #     payment = payments.filter(month_period=month_period).first()
            #     if payment:
            #         all_payments.append({
            #             'month_period': month_period,
            #             'amount': payment.amount,
            #             'status': payment.status,
            #             'payment_method': payment.payment_method,
            #             'date': payment.date
            #         })
            #     else:
            #         all_payments.append({
            #             'month_period': month_period,
            #             'amount': student.group.course.price_per_month,
            #             'status': 'unpaid',
            #             'payment_method': None,
            #             'date': None
            #         })
            duration = student.group.course.duration_months
            # payments = Payment.objects.filter(
            #     student=student,
            #     group=student.group
            # ).order_by('month_period')

            payments_dict = {payment.month_period: payment for payment in payments}
            all_payments = []
            monthly_payment = student.group.course.price_per_month

            for month_period in MonthPeriod.objects.filter(is_active=True):
                if month_period in payments_dict:
                    payment = payments_dict[month_period]
                    all_payments.append({
                        'month': month_period,
                        'amount': payment.amount,
                        'status': payment.status,
                        'payment_method': payment.payment_method,
                        'date': payment.date
                    })
                else:
                    all_payments.append({
                        'month': month_period,
                        'amount': student.group.course.price_per_month,
                        'status': 'unpaid',
                        'payment_method': None,
                        'date': None
                    })
            context['all_payments'] = all_payments
        return context


class StudentProfileForTeacherView(LoginRequiredMixin, DetailView):
    model = User
    template_name = 'teacher/student_profile_for_teacher.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object

        if student.group:
            context['course'] = student.group.course
            context['start_date'] = student.group.start_date
            context['end_date'] = student.group.end_date

            # Hozirgi o'qiyotgan oyi
            start_date = student.group.start_date
            today = date.today()
            months_studied = relativedelta(today, start_date).months + 1
            context['current_month'] = min(months_studied, student.group.course.duration_months)

            # Kurs davomiyligi bo'yicha to'lovlar ro'yxati
            duration = student.group.course.duration_months
            payments = Payment.objects.filter(
                student=student,
                group=student.group
            )

            # To'lovlar lug'atini yaratish
            payments_dict = {payment.month_period: payment for payment in payments}

            # Barcha oylar uchun to'lovlar ro'yxati
            all_payments = []
            monthly_payment = student.group.course.price_per_month

            for month_period in range(1, duration + 1):
                if month_period in payments_dict:
                    payment = payments_dict[month_period]
                    all_payments.append({
                        'month': month_period,
                        'amount': payment.amount,
                        'date': payment.date,
                        'status': payment.status,
                        'comment': payment.comment
                    })
                else:
                    all_payments.append({
                        'month': month_period,
                        'amount': monthly_payment,
                        'date': None,
                        'status': 'unpaid',
                        'comment': None
                    })

            context['all_payments'] = all_payments

            # Umumiy to'langan summa
            total_paid = payments.filter(
                status='paid'
            ).aggregate(total=Sum('amount'))['total'] or 0
            context['total_paid'] = total_paid

            # Baholar
            context['grades'] = Grade.objects.filter(student=student).order_by('-date')

            # O'rtacha baho
            if context['grades'].exists():
                avg_grade = sum(grade.grade for grade in context['grades']) / context['grades'].count()
                context['average_grade'] = round(avg_grade, 1)

        return context


class PaymentPersonalView(LoginRequiredMixin, View):
    def post(self, request, student_id, period_id):
        student = get_object_or_404(User, id=student_id)
        period = get_object_or_404(MonthPeriod, id=period_id)

        if not student.group:
            messages.error(request, "O'quvchi hech qaysi guruhga biriktirilmagan!")
            return redirect('student_profile', pk=student_id)

        form = PaymentPersonalForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.student = student
            payment.group = student.group
            payment.course = student.group.course
            payment.month_period = period
            payment.status = 'paid'
            try:
                payment.full_clean()
                payment.save()
                messages.success(request, f"{period.name} uchun to'lov muvaffaqiyatli amalga oshirildi!")
            except ValidationError as e:
                messages.error(request, f"Validatsiya xatosi: {e}")
        else:
            messages.error(request, f"Form xatosi: {form.errors}")
        return redirect('student_profile', pk=student_id)


class PaymentHistoryView(LoginRequiredMixin, DetailView):
    model = User
    template_name = 'payments/history.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object

        # To'lovlar statistikasi
        stats = Payment.objects.filter(student=student).aggregate(
            total_paid=Sum('amount', filter=models.Q(status='paid')),
            total_payments=Count('id', filter=models.Q(status='paid')),
            avg_payment=models.Avg('amount', filter=models.Q(status='paid'))
        )

        context.update({
            'payments': Payment.objects.filter(student=student).order_by('-date'),
            'stats': stats
        })

        return context


class PaymentUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Payment
    form_class = PaymentPersonalForm
    template_name = 'payments/update.html'

    def test_func(self):
        # Faqat manager va teacher lar to'lovni tahrirlashi mumkin
        return self.request.user.role in ['manager', 'teacher']

    def get_success_url(self):
        return reverse('student_profile', kwargs={'pk': self.object.student.id})


class PaymentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Payment
    template_name = 'payments/delete.html'

    def test_func(self):
        # Faqat manager va teacher lar to'lovni o'chirishi mumkin
        return self.request.user.role in ['manager', 'teacher']

    def get_success_url(self):
        return reverse('student_profile', kwargs={'pk': self.object.student.id})


def is_teacher(user):
    return user.role == 'teacher'


@login_required
@user_passes_test(is_teacher)
def teacher_profile(request):
    teacher = request.user

    # Guruhlar va o'quvchilar
    groups = Group.objects.filter(teacher=teacher)
    students_count = User.objects.filter(group__teacher=teacher, role='student').count()

    # To'lov statistikalari
    payments = Payment.objects.filter(group__teacher=teacher)
    total_paid_amount = payments.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    total_unpaid_count = payments.filter(status='unpaid').values('student').distinct().count()
    paid_students_count = Payment.objects.filter(group__teacher=teacher, status='paid').values('student').distinct().count()
    unpaid_students_count = students_count - paid_students_count

    context = {
        'groups': groups,
        'total_paid_amount': total_paid_amount,
        'total_unpaid_count': total_unpaid_count,
        'students_count': students_count,
        'paid_students_count': paid_students_count,
        'unpaid_students_count': unpaid_students_count,
        'schedule': groups
    }

    return render(request, 'teacher/profile.html', context)


@login_required
@user_passes_test(is_teacher)
def unpaid_students(request):
    teacher = request.user

    # Teacher guruhidagi barcha o'quvchilar
    all_students = User.objects.filter(group__teacher=teacher, role='student')

    # To'lov qilgan talabalar
    paid_students_ids = Payment.objects.filter(group__teacher=teacher, status='paid').values_list('student_id', flat=True)

    # To'lov qilmagan talabalar
    unpaid_students = all_students.exclude(id__in=paid_students_ids).values(
        'id', 'first_name', 'last_name', 'phone', 'group__name'
    ).distinct()

    return JsonResponse(list(unpaid_students), safe=False)



@login_required
@user_passes_test(is_teacher)
def teacher_groups(request):
    teacher = request.user
    groups = Group.objects.filter(teacher=teacher)
    return render(request, 'teacher/groups.html', {'groups': groups})



@login_required
@user_passes_test(is_teacher)
def teacher_students(request):
    teacher = request.user
    today = datetime.today()
    students = User.objects.filter(group__teacher=teacher, role='student').annotate(
        total_payments=Sum('payment__amount'),
        course_months=F('group__course__duration_months'),
        course_start_date=F('group__start_date'),
        # monthly_price=F('group__course__price_per_month')
    )

    student_data = []
    for student in students:
        active_periods = MonthPeriod.objects.filter(
            start_date__gte=student.group.start_date,
            end_date__lte=student.group.end_date
        )
        total_paid = student.total_payments or 0
        required_payment = sum([330000 for period in active_periods])
        payment_status = "Paid" if total_paid >= required_payment else "Unpaid"

        student_data.append({
            'pk': student.id,
            'username': student.username,
            'full_name': f"{student.first_name} {student.last_name}",
            'group_name': student.group.name,
            'total_payments': total_paid,
            'required_payment': required_payment,
            'payment_status': payment_status,
        })
    return render(request, 'teacher/students.html', {'students': student_data})




@login_required
@user_passes_test(is_teacher)
def add_grade_profile(request):
    if request.method == 'POST':
        form = GradeForm(request.POST)
        if form.is_valid():
            grade = form.save(commit=False)
            grade.teacher = request.user
            grade.save()
            messages.success(request, "Baho muvaffaqiyatli qo'shildi")
            return redirect('teacher_students')
    else:
        form = GradeForm()

    teacher = request.user
    students = User.objects.filter(group__teacher=teacher, role='student')
    context = {
        'form': form,
        'students': students,
        'recent_grades': Grade.objects.filter(teacher=teacher).order_by('-date')[:10]
    }
    return render(request, 'teacher/add_grade.html', context)

@login_required
@user_passes_test(is_teacher)
def get_group_students(request, group_id):
    group = get_object_or_404(Group, id=group_id, teacher=request.user)
    students = group.users.filter(role='student').values(
        'id', 'first_name', 'last_name', 'phone', 'image'
    ).annotate(
        average_grade=Avg('grades__grade'),
        full_name=Concat('first_name', Value(' '), 'last_name')
    )


from django.db.models import Avg, Value
from django.db.models.functions import Concat
from django.utils.timezone import now

from django.templatetags.static import static
from django.conf import settings
from urllib.parse import urljoin

@login_required
@user_passes_test(is_teacher)
def get_group_students(request, group_id):
    group = get_object_or_404(Group, id=group_id, teacher=request.user)
    today = now().date()  # Bugungi sana

    students = group.users.filter(role='student').annotate(
        sum=Sum('grades__grade'),
        average_grade=Avg('grades__grade'),
        full_name=Concat('first_name', Value(' '), 'last_name'),
        today_grade=Subquery(
            Grade.objects.filter(
                student=OuterRef('id'),
                date=today
            ).values('grade')[:1]
        )
    ).values(
        'id', 'full_name', 'phone', 'image', 'sum', 'average_grade', 'today_grade'
    )

    # Har bir studentning rasmiga to'liq URL qo'shish
    students_with_images = []
    for student in students:
        if student['image']:
            student['image'] = urljoin(settings.MEDIA_URL, student['image'])  # Media URL bilan birlashtirish
        else:
            student['image'] = static('default_profile.png')  # Agar rasm bo'lmasa, default rasm
        students_with_images.append(student)

    return JsonResponse(students_with_images, safe=False)



@csrf_exempt
@login_required
@user_passes_test(is_teacher)
def add_grade(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('student')
            grade_value = data.get('grade')

            if not student_id or not grade_value:
                return JsonResponse({
                    'success': False,
                    'error': 'Student ID va baho talab qilinadi'
                }, status=400)

            student = User.objects.get(id=student_id, role='student')

            # Bugungi sanani olish
            today = timezone.now().date()

            # Bugungi kunga mavjud bahoni tekshirish
            existing_grade = Grade.objects.filter(
                student=student,
                teacher=request.user,
                date=today  # Baholash bugungi sanaga mos bo'lishi kerak
            ).first()

            if existing_grade:
                # Agar bugungi kunga baho bo'lsa, uni yangilaymiz
                existing_grade.grade = grade_value
                existing_grade.save()
            else:
                # Agar baho yo'q bo'lsa, yangi baho qo'shamiz
                Grade.objects.create(
                    student=student,
                    teacher=request.user,
                    grade=grade_value,
                    date=today
                )

            # O'rtacha bahoni hisoblash
            avg_grade = Grade.objects.filter(
                student=student
            ).aggregate(Avg('grade'))['grade__avg']

            return JsonResponse({
                'success': True,
                'average_grade': avg_grade,
                'message': 'Baho muvaffaqiyatli qo\'shildi yoki yangilandi'
            })

        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'O\'quvchi topilmadi'
            }, status=404)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

    return JsonResponse({
        'success': False,
        'error': 'Faqat POST so\'rovlar qabul qilinadi'
    }, status=405)


class PaymentReportView(TemplateView):
    template_name = 'payments/report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get('search', '')

        # Tanlangan oy (GET parametri orqali)
        selected_period_id = self.request.GET.get('period')
        if selected_period_id:
            selected_period = MonthPeriod.objects.get(id=selected_period_id)
            payments = Payment.objects.filter(month_period=selected_period, status='paid')
        else:
            selected_period = None
            payments = Payment.objects.filter(status='paid')

        # Umumiy statistika
        total_payments = payments.count()
        total_amount = payments.aggregate(Sum('amount'))['amount__sum'] or 0

        # To'lov turlari bo'yicha statistika
        payment_methods_stats = payments.values('payment_method').annotate(
            count=Count('id'),
            total=Sum('amount')
        )
        if search_query:
            payments = payments.filter(
                Q(student__first_name__icontains=search_query) |
                Q(student__last_name__icontains=search_query) |
                Q(group__name__icontains=search_query)
            )
        context.update({
            'periods': MonthPeriod.objects.all(),
            'selected_period': selected_period,
            'payments': payments.select_related('student', 'group'),
            'total_payments': total_payments,
            'total_amount': total_amount,
            'payment_methods_stats': payment_methods_stats,
        })
        return context


class PaymentSearchView(ListView):
    model = Payment
    template_name = 'payments/_payments_table.html'
    context_object_name = 'payments'

    def get_queryset(self):
        queryset = Payment.objects.filter(status='paid')  # Faqat to'langanlarni olish
        search = self.request.GET.get('q', '')
        if search:
            queryset = queryset.filter(
                Q(student__first_name__icontains=search) |
                Q(student__last_name__icontains=search) |
                Q(group__name__icontains=search) |
                Q(amount__icontains=search)  # Summa bo'yicha qidirish
            )
        return queryset

class PaymentCreateView(View):
    template_name = 'payments/payment_form.html'

    def get(self, request):
        initial_data = {}
        group_id = request.GET.get('group')
        if group_id:
            try:
                group = Group.objects.get(id=group_id)
                initial_data = {
                    'group': group,
                    'amount': group.course.price_per_month,
                    'course': group.course,

                }
            except Group.DoesNotExist:
                pass

        form = PaymentForm(initial=initial_data)
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = PaymentForm(request.POST)
        if form.is_valid():
            try:
                payment = form.save(commit=False)
                payment.course = payment.group.course
                payment.status = 'paid'
                payment.save()
                messages.success(request, "To'lov muvaffaqiyatli qo'shildi!")
                return HttpResponse("""
                <div class="alert alert-success">
                    To'lov muvaffaqiyatli qo'shildi!
                </div>
            """)
            except Exception as e:
                messages.error(request, f"Xatolik yuz berdi: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
        return render(request, self.template_name, {'form': form})


# class PaymentCreateView(View):
#     template_name = 'payments/payment_form.html'
#
#     def get(self, request):
#         form = PaymentForm()
#         return render(request, self.template_name, {'form': form})
#
#     def post(self, request):
#         form = PaymentForm(request.POST)
#         if form.is_valid():
#             payment = form.save()
#             messages.success(request, "To'lov muvaffaqiyatli qo'shildi!")
#             return HttpResponse("<div class='alert alert-success'>To'lov muvaffaqiyatli qo'shildi!</div>")
#
#         return render(request, self.template_name, {'form': form})

# class PaymentCreateView(View):
#     template_name = 'payments/payment_form.html'
#
#     def get(self, request):
#         form = PaymentForm()
#         return render(request, self.template_name, {'form': form})
#
#     def post(self, request):
#         form = PaymentForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return HttpResponse("""
#                 <div class="alert alert-success">
#                     To'lov muvaffaqiyatli qo'shildi!
#                 </div>
#                 <script>
#                     window.setTimeout(() => {
#                         window.location.reload();
#                     }, 1500);
#                 </script>
#             """)
#         return render(request, self.template_name, {'form': form})